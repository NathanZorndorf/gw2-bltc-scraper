import os
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import matplotlib.pyplot as plt
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# --------------------
# NEW: TRANSACTION HISTORY SCRAPER
# --------------------
def parse_coins_to_gold_silver(coins):
    # API returns price in coins (1g = 10000, 1s = 100)
    gold = coins // 10000
    silver = (coins % 10000) // 100
    return round(gold + silver / 100, 2)

def fetch_all_transactions(endpoint, api_key):
    headers = {"Authorization": f"Bearer {api_key}"}
    all_tx = []
    page = 0
    page_size = 200  # max allowed by API
    while True:
        url = f"{endpoint}?page={page}&page_size={page_size}"
        r = requests.get(url, headers=headers, timeout=20)
        # Print the URL
        print(f"URL: {r.request.url}")

        # Print the headers
        print("Headers:")
        for header, value in r.request.headers.items():
            print(f"  {header}: {value}")

        if r.status_code != 200:
            break
        batch = r.json()
        if not batch:
            break
        all_tx.extend(batch)
        if len(batch) < page_size:
            break
        page += 1
    return all_tx

def fetch_transactions(api_key):
    base = "https://api.guildwars2.com/v2/commerce/transactions/history"
    buys = fetch_all_transactions(f"{base}/buys", api_key)
    sells = fetch_all_transactions(f"{base}/sells", api_key)
    return buys, sells

def get_item_names(item_ids):
    # Fetch item names from API in batches of 200
    names = {}
    ids = list(set(item_ids))
    for i in range(0, len(ids), 200):
        batch = ids[i:i+200]
        try:
            r = requests.get(
                "https://api.guildwars2.com/v2/items",
                params={"ids": ",".join(map(str, batch))},
                timeout=20
            )
            r.raise_for_status()
            for item in r.json():
                # Defensive: some items may be missing or malformed
                if isinstance(item, dict) and "id" in item and "name" in item:
                    names[item["id"]] = item["name"]
        except Exception as e:
            print(f"Error fetching item names for batch {batch}: {e}")
    # Fallback for missing names
    for iid in ids:
        if iid not in names:
            names[iid] = f"Item {iid}"
    return names

def filter_last_n_days(transactions, date_field="purchased", n=30):
    cutoff = datetime.now() - timedelta(days=n)
    filtered = []
    skipped = 0
    for tx in transactions:
        date_str = tx.get(date_field) or tx.get("created")
        if not date_str:
            skipped += 1
            continue
        try:
            # Handle both Z and +00:00
            if date_str.endswith("Z"):
                tx_date = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S%z")
            else:
                # Remove colon in timezone if present (Python <3.7 compatibility)
                if "+" in date_str:
                    main, tz = date_str.split("+")
                    tz = tz.replace(":", "")
                    date_str = main + "+" + tz
                tx_date = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S%z")
            if tx_date >= cutoff.replace(tzinfo=tx_date.tzinfo):
                filtered.append(tx)
        except Exception as e:
            skipped += 1
            continue
    print(f"Total transactions: {len(transactions)}, after filtering: {len(filtered)}, skipped: {skipped}")
    return filtered

def aggregate_transactions(buys, sells):
    # Only require at least one buy in the last 7 days
    agg = {}
    for tx in buys:
        iid = tx["item_id"]
        price = parse_coins_to_gold_silver(tx["price"])
        qty = tx["quantity"]
        spent = price * qty
        if iid not in agg:
            agg[iid] = {"bought_qty": 0, "spent": 0, "sold_qty": 0, "received": 0}
        agg[iid]["bought_qty"] += qty
        agg[iid]["spent"] += spent
    for tx in sells:
        iid = tx["item_id"]
        price = parse_coins_to_gold_silver(tx["price"])
        qty = tx["quantity"]
        received = price * qty * 0.85
        if iid not in agg:
            continue  # Only include items with buys
        agg[iid]["sold_qty"] += qty
        agg[iid]["received"] += received
    # Only keep items with at least one buy
    return {iid: data for iid, data in agg.items() if data["bought_qty"] > 0}

def save_profit_report(agg, item_names, output_file):
    rows = []
    for iid, data in agg.items():
        name = item_names.get(iid, f"Item {iid}")
        spent = data["spent"]
        received = data["received"]
        roi = received - spent
        roi_pct = roi / spent if spent else ""
        sum_qty = data["bought_qty"] + data["sold_qty"]
        rows.append({
            "Item Name": name,
            "Bought Qty": data["bought_qty"],
            "Sold Qty": data["sold_qty"],
            "Sum Qty": sum_qty,
            "Total Spent (g.s)": spent,
            "Total Received (g.s)": received,
            "ROI (g.s)": roi,
            "ROI (%)": roi_pct
        })
    df = pd.DataFrame(rows)
    if df.empty or "ROI (g.s)" not in df.columns:
        print("No transactions to report for the selected period/items.")
        return
    # Ensure ROI (%) is numeric for sorting/plotting
    df["ROI (%)"] = pd.to_numeric(df["ROI (%)"], errors="coerce").fillna(0)
    df = df.sort_values("ROI (g.s)", ascending=False)
    df.to_excel(output_file, index=False)

    # Format Excel
    wb = load_workbook(output_file)
    ws = wb.active
    ws.freeze_panes = 'B2'
    header_to_idx = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1)}
    def L(name): return get_column_letter(header_to_idx[name])
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        ws[f'{L("Total Spent (g.s)")}{row}'].number_format = '0.00'
        ws[f'{L("Total Received (g.s)")}{row}'].number_format = '0.00'
        ws[f'{L("ROI (g.s)")}{row}'].number_format = '0.00'
        ws[f'{L("ROI (%)")}{row}'].number_format = '0.00%'
        ws[f'{L("Bought Qty")}{row}'].number_format = '0'
        ws[f'{L("Sold Qty")}{row}'].number_format = '0'
        ws[f'{L("Sum Qty")}{row}'].number_format = '0'
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(8, max_length + 2)
    wb.save(output_file)
    print(f"Profit report saved to {output_file}")

    # --- Analytics Visualizations ---
    print("Generating analytics visualizations...")

    # Ensure img/ folder exists
    img_dir = os.path.join(os.path.dirname(output_file), "img")
    os.makedirs(img_dir, exist_ok=True)

    # Pie chart: Profits by item
    top_profit_items = df[df["ROI (g.s)"] > 0].sort_values("ROI (g.s)", ascending=False).head(10)
    plt.figure(figsize=(8,8))
    plt.pie(top_profit_items["ROI (g.s)"], labels=top_profit_items["Item Name"], autopct='%1.1f%%', startangle=140)
    plt.title("Top 10 Profitable Items (ROI in Gold)")
    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, "profit_pie_chart.png"))
    plt.close()

    # Bar chart: ROI (%) by item
    top_roi_pct = df.sort_values("ROI (%)", ascending=False).head(10)
    plt.figure(figsize=(10,6))
    plt.bar(top_roi_pct["Item Name"], top_roi_pct["ROI (%)"])
    plt.xticks(rotation=45, ha='right')
    plt.title("Top 10 Items by ROI (%)")
    plt.ylabel("ROI (%)")
    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, "roi_percent_bar_chart.png"))
    plt.close()

    # Bar chart: Highest profit items (ROI (g.s))
    top_profit = df.sort_values("ROI (g.s)", ascending=False).head(10)
    plt.figure(figsize=(10,6))
    plt.bar(top_profit["Item Name"], top_profit["ROI (g.s)"])
    plt.xticks(rotation=45, ha='right')
    plt.title("Top 10 Highest Profit Items (Gold)")
    plt.ylabel("Profit (g.s)")
    plt.tight_layout()
    plt.savefig(os.path.join(img_dir, "top_profit_items.png"))
    plt.close()

    print("Analytics charts saved to img/: profit_pie_chart.png, roi_percent_bar_chart.png, top_profit_items.png")

# --------------------
# MAIN ENTRY FOR PROFIT REPORT
# --------------------
if __name__ == "__main__":
    api_key = os.environ.get("GW2_API_KEY")
    if not api_key:
        print("Error: GW2_API_KEY environment variable not set. Please set it before running this script.")
    else:
        print("Fetching transaction history from API...")
        buys, sells = fetch_transactions(api_key)
        buys = filter_last_n_days(buys, date_field="purchased", n=30)
        sells = filter_last_n_days(sells, date_field="purchased", n=30)
        all_ids = [tx["item_id"] for tx in buys + sells]
        item_names = get_item_names(all_ids)
        agg = aggregate_transactions(buys, sells)
        PROFIT_FILE = "profit-report.xlsx"
        save_profit_report(agg, item_names, PROFIT_FILE)
        save_profit_report(agg, item_names, PROFIT_FILE)
