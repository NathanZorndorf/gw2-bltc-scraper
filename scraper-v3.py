import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

BASE_URL = "https://www.gw2bltc.com/en/tp/search"
PARAMS = {
    "profit-min": 5000,
    "profit-pct-min": 10,
    "profit-pct-max": 100,
    "sold-day-min": 20,
    "bought-day-min": 20,
    "ipg": 200,
    "sort": "profit-pct",
    "page": 1
}

OVERCUT_PCT = 1.1
UNDERCUT_PCT = 0.9

scrape_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
all_data = []

def parse_gold_silver(td):
    gold = silver = 0
    for span in td.find_all("span"):
        classes = span.get("class", [])
        if "cur-t1c" in classes:
            gold = int(span.get_text(strip=True))
        elif "cur-t1b" in classes:
            silver = int(span.get_text(strip=True))
    return round(gold + silver / 100, 2)

def parse_int(td):
    return int(td.get_text(strip=True).replace(",", ""))

while True:
    print(f"Fetching page {PARAMS['page']}...")
    r = requests.get(BASE_URL, params=PARAMS)
    if r.status_code != 200:
        break

    soup = BeautifulSoup(r.text, "html.parser")
    rows = soup.select("table.table-result tr")[1:]  

    if not rows:
        break

    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 12:
            continue

        item_name = cols[1].get_text(strip=True)
        link_tag = cols[1].find("a", href=True)
        item_link = f"https://www.gw2bltc.com{link_tag['href']}" if link_tag else ""

        sell = parse_gold_silver(cols[2])
        buy = parse_gold_silver(cols[3])
        supply = parse_int(cols[6])
        demand = parse_int(cols[7])
        sold = parse_int(cols[8])
        offers = parse_int(cols[9])
        bought = parse_int(cols[10])
        bids = parse_int(cols[11])

        all_data.append([
            item_name, item_link, scrape_date, buy, sell, 
            demand, supply, offers, sold, bids, bought
        ])

    PARAMS["page"] += 1

# Create DataFrame
df = pd.DataFrame(all_data, columns=[
    "Item Name", "Item Link", "Date of Scrape",
    "Buy (g.s)", "Sell (g.s)", "Demand", "Supply", 
    "Offers", "Sold", "Bids", "Bought"
])

# Add static % columns
df["Overcut (%)"] = OVERCUT_PCT
df["Undercut (%)"] = UNDERCUT_PCT

# Boolean columns for checkboxes
df["Order Placed"] = False
df["Order Successful"] = False
df["Sold?"] = False

# Save initial DataFrame to Excel
output_file = "gw2_trading_post.xlsx"
df.to_excel(output_file, index=False)

# Add formulas with openpyxl
wb = load_workbook(output_file)
ws = wb.active
max_row = ws.max_row

# Insert formulas starting from row 2
for row in range(2, max_row + 1):
    ws[f"N{row}"] = f"=E{row}*J{1}"  # Overcut (g) formula (Buy * Overcut %)
    ws[f"O{row}"] = f"=D{row}*K{1}"  # Undercut (g) formula (Sell * Undercut %)
    ws[f"P{row}"] = f"=O{row}*0.85 - N{row}"  # Theoretical Profit
    ws[f"Q{row}"] = f"=O{row}*0.85"           # Amount Received

wb.save(output_file)
print(f"Excel file with formulas saved as {output_file}")
