import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --------------------
# CONFIG
# --------------------
BASE_URL = "https://www.gw2bltc.com/en/tp/search"
PARAMS = {
    "profit-min": 5000,
    "profit-pct-min": 10,
    "profit-pct-max": 100,
    "sold-day-min": 20,
    "bought-day-min": 20,
    "ipg": 200,
    "sort": "profit-pct",
    "page": 1
}

OVERCUT_PCT = 1.10
UNDERCUT_PCT = 0.90

scrape_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
all_data = []

# --------------------
# PARSERS
# --------------------
def parse_gold_silver(td):
    gold = silver = 0
    for span in td.find_all("span"):
        classes = span.get("class", [])
        if "cur-t1c" in classes:
            gold = int(span.get_text(strip=True))
        elif "cur-t1b" in classes:
            silver = int(span.get_text(strip=True))
    return round(gold + silver / 100, 2)

def parse_int(td):
    return int(td.get_text(strip=True).replace(",", ""))

# --------------------
# SCRAPING LOOP
# --------------------
while True:
    print(f"Fetching page {PARAMS['page']}...")
    r = requests.get(BASE_URL, params=PARAMS)
    if r.status_code != 200:
        print("Request failed.")
        break

    soup = BeautifulSoup(r.text, "html.parser")
    rows = soup.select("table.table-result tr")[1:]  # skip header

    if not rows:
        break

    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 12:
            continue

        item_name = cols[1].get_text(strip=True)
        link_tag = cols[1].find("a", href=True)
        item_link = f"https://www.gw2bltc.com{link_tag['href']}" if link_tag else ""

        sell = parse_gold_silver(cols[2])
        buy = parse_gold_silver(cols[3])
        supply = parse_int(cols[6])
        demand = parse_int(cols[7])
        sold = parse_int(cols[8])
        offers = parse_int(cols[9])
        bought = parse_int(cols[10])
        bids = parse_int(cols[11])

        all_data.append([
            item_name, item_link, scrape_date, sell, buy,
            supply, demand, sold, offers, bought, bids
        ])

    PARAMS["page"] += 1

# --------------------
# CREATE BASE DATAFRAME
# --------------------
df = pd.DataFrame(all_data, columns=[
    "Item Name", "Item Link", "Date of Scrape",
    "Sell (g.s)", "Buy (g.s)", "Supply", "Demand", "Sold",
    "Offers", "Bought", "Bids"
])

df["Overcut (%)"] = OVERCUT_PCT
df["Undercut (%)"] = UNDERCUT_PCT

# Boolean placeholders for checkboxes
df["Order Placed"] = False
df["Order Successful"] = False
df["Sold?"] = False

output_file = "gw2_trading_post.xlsx"
df.to_excel(output_file, index=False)

# --------------------
# ADD FORMULAS & FORMATTING
# --------------------
wb = load_workbook(output_file)
ws = wb.active
max_row = ws.max_row

# Find column indices for formulas
col_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}

for row in range(2, max_row + 1):
    ws.cell(row=row, column=col_map["Overcut (%)"]).number_format = "0.00"
    ws.cell(row=row, column=col_map["Undercut (%)"]).number_format = "0.00"

    ws.cell(row=row, column=col_map["Sell (g.s)"]).number_format = "0.00"
    ws.cell(row=row, column=col_map["Buy (g.s)"]).number_format = "0.00"

    # Overcut (g) = Buy * Overcut %
    ws.cell(row=row, column=col_map["Overcut (g)"]+1, value=f"=E{row}*J{row}").number_format = "0.00"
    # Undercut (g) = Sell * Undercut %
    ws.cell(row=row, column=col_map["Undercut (g)"]+2, value=f"=D{row}*K{row}").number_format = "0.00"
    # Theoretical Profit
    ws.cell(row=row, column=col_map["Theoretical Profit"]+3, value=f"=O{row}*0.85 - N{row}").number_format = "0.00"
    # Amount Received
    ws.cell(row=row, column=col_map["Amount Received"]+4, value=f"=O{row}*0.85").number_format = "0.00"
    # Demand-Supply Gap (%)
    ws.cell(row=row, column=col_map["Demand-Supply Gap (%)"]+5, value=f"=(G{row}-F{row})/F{row}").number_format = "0.00%"

# --------------------
# AUTO-FIT COLUMN WIDTHS
# --------------------
for col in ws.columns:
    if col==col_map["Item Link"]:
        continue
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value is not None:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
    ws.column_dimensions[col_letter].width = max_length + 1

wb.save(output_file)
print(f"Excel file saved as {output_file} with {len(df)} rows and formulas.")
