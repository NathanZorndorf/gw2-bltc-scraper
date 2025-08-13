import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

# --------------------
# CONFIG
# --------------------
BASE_URL = "https://www.gw2bltc.com/en/tp/search"
PARAMS = {
    "profit-min": 500,
    "profit-pct-min": 0,
    "profit-pct-max": 100,
    "sold-day-min": 10,
    "bought-day-min": 10,
    "ipg": 200,
    "sort": "profit-pct",
    "page": 1
}

OVERCUT_PCT_DEFAULT = 1.10
UNDERCUT_PCT_DEFAULT = 0.90

output_file = "gw2_trading_post_sorted.xlsx"

# --------------------
# HELPER FUNCTIONS
# --------------------
def parse_gold_silver(td):
    """Helper function to parse gold and silver values from table cells."""
    gold = silver = 0
    for span in td.find_all("span"):
        classes = span.get("class", [])
        if "cur-t1c" in classes:
            try:
                gold = int(span.get_text(strip=True))
            except ValueError:
                gold = 0
        elif "cur-t1b" in classes:
            try:
                silver = int(span.get_text(strip=True))
            except ValueError:
                silver = 0
    return round(gold + silver / 100, 2)

def parse_int(td):
    """Helper function to parse integer values from table cells."""
    txt = td.get_text(strip=True).replace(",", "")
    try:
        return int(txt)
    except ValueError:
        return 0

def get_total_pages(soup):
    """Finds the total number of pages from the pagination element."""
    # Look for the pagination buttons
    pagination_buttons = soup.select('.btn-group.btn-group-justified a.btn')
    
    if not pagination_buttons:
        print("No pagination buttons found")
        return 1
    
    max_page = 1
    
    # Look through all pagination links to find the highest page number
    for button in pagination_buttons:
        href = button.get('href', '')
        if 'page=' in href:
            # Extract page number from href
            page_match = re.search(r'page=(\d+)', href)
            if page_match:
                page_num = int(page_match.group(1))
                max_page = max(max_page, page_num)
        
        # Also check button text for page numbers
        button_text = button.get_text(strip=True)
        if button_text.isdigit():
            page_num = int(button_text)
            max_page = max(max_page, page_num)
    
    print(f"Detected pagination: Found max page {max_page}")
    return max_page

# --------------------
# PRE-SCRAPE CHECK
# --------------------
print("Checking total number of pages...")
try:
    first_page_r = requests.get(BASE_URL, params=PARAMS, timeout=20)
    first_page_r.raise_for_status()
    first_page_soup = BeautifulSoup(first_page_r.text, "html.parser")
    
    # Debug: Print pagination HTML to see what we're working with
    pagination_div = first_page_soup.select_one('.btn-group.btn-group-justified')
    if pagination_div:
        print("Found pagination HTML:")
        print(pagination_div.prettify()[:500] + "...")
    else:
        print("No pagination div found")
    
    total_pages = get_total_pages(first_page_soup)
    
    print(f"Found {total_pages} pages of results.")
    
    # If more than 5 pages, ask for confirmation
    if total_pages > 5:
        user_choice = input(f"Do you want to scrape all {total_pages} pages? (y/n): ").lower()
        if user_choice != 'y':
            print("Scraping cancelled by user.")
            exit()
    else:
        print(f"Proceeding with {total_pages} pages...")

except requests.exceptions.RequestException as e:
    print(f"Failed to fetch initial page data: {e}")
    exit()

# --------------------
# SCRAPE -> list of rows
# --------------------
all_rows = []
scrape_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Loop through pages on the website until there's no more data
current_page = 1
while current_page <= total_pages:
    PARAMS['page'] = current_page
    print(f"Fetching page {current_page}/{total_pages}...")
    try:
        r = requests.get(BASE_URL, params=PARAMS, timeout=20)
        r.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Request failed on page {current_page}: {e}")
        break # Stop if a page fails

    soup = BeautifulSoup(r.text, "html.parser")
    rows = soup.select("table.table-result tr")[1:]

    if not rows:
        print(f"No more rows found on page {current_page}. Ending scrape.")
        break

    print(f"Found {len(rows)} items on page {current_page}")

    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 12:
            continue

        item_name = cols[1].get_text(strip=True)
        link_tag = cols[1].find("a", href=True)
        item_link = f"https://www.gw2bltc.com{link_tag['href']}" if link_tag else ""

        sell = parse_gold_silver(cols[2])
        buy = parse_gold_silver(cols[3])
        supply = parse_int(cols[6])
        demand = parse_int(cols[7])
        sold = parse_int(cols[8])
        offers = parse_int(cols[9])
        bought = parse_int(cols[10])
        bids = parse_int(cols[11])

        all_rows.append([
            item_name, item_link, scrape_time, sell, buy,
            supply, demand, sold, offers, bought, bids
        ])
    
    current_page += 1

if not all_rows:
    print("No data was scraped. Exiting.")
    exit()

print(f"Successfully scraped {len(all_rows)} total items.")

# --------------------
# PROCESS DATA with Pandas
# --------------------
print("Processing data...")
df = pd.DataFrame(all_rows, columns=[
    "Item Name", "Item Link", "Date of Scrape",
    "Sell (g.s)", "Buy (g.s)", "Supply", "Demand", "Sold",
    "Offers", "Bought", "Bids"
])

# Add placeholder columns for formulas (will be filled with actual formulas in Excel)
df["Overcut (%)"] = OVERCUT_PCT_DEFAULT
df["Undercut (%)"] = UNDERCUT_PCT_DEFAULT
df["Overcut (g)"] = 0  # Will be replaced with formula
df["Undercut (g)"] = 0  # Will be replaced with formula
df["Theoretical Profit"] = 0  # Will be replaced with formula
df["Amount Received"] = 0  # Will be replaced with formula
df["Demand-Supply Gap (%)"] = 0  # Will be replaced with formula

# Add manual columns
df["Order Placed"] = False
df["Order Successful"] = False
df["Sold (manual)"] = False

# Sort data by profit
df = df.sort_values(by="Theoretical Profit", ascending=False)

# Reorder columns to the desired final layout
final_column_order = [
    "Item Name", "Item Link", "Date of Scrape", "Buy (g.s)", "Sell (g.s)",
    "Supply", "Demand", "Offers", "Sold", "Bids", "Bought",
    "Overcut (%)", "Undercut (%)", "Overcut (g)", "Undercut (g)", 
    "Theoretical Profit", "Amount Received", "Demand-Supply Gap (%)",
    "Order Placed", "Order Successful", "Sold (manual)"
]
df = df[final_column_order]

# Save the reordered and sorted dataframe to Excel.
df.to_excel(output_file, index=False)
print(f"Data saved to {output_file}. Now applying final formatting...")

# --------------------
# FORMAT EXCEL FILE with OpenPyXL
# --------------------
wb = load_workbook(output_file)
ws = wb.active

# Freeze the top row
ws.freeze_panes = 'A2'

max_row = ws.max_row
header_to_idx = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1)}

def L(name):
    if name not in header_to_idx:
        raise KeyError(f"Column '{name}' not found in the Excel sheet headers.")
    return get_column_letter(header_to_idx[name])

# Apply number formats and formulas
for row in range(2, max_row + 1):
    # Number formats
    ws[f'{L("Buy (g.s)")}{row}'].number_format = '0.00'
    ws[f'{L("Sell (g.s)")}{row}'].number_format = '0.00'
    ws[f'{L("Overcut (%)")}{row}'].number_format = '0.00%'
    ws[f'{L("Undercut (%)")}{row}'].number_format = '0.00%'
    ws[f'{L("Overcut (g)")}{row}'].number_format = '0.00'
    ws[f'{L("Undercut (g)")}{row}'].number_format = '0.00'
    ws[f'{L("Theoretical Profit")}{row}'].number_format = '0.00'
    ws[f'{L("Amount Received")}{row}'].number_format = '0.00'
    ws[f'{L("Demand-Supply Gap (%)")}{row}'].number_format = '0.00%'
    for int_col in ["Supply", "Demand", "Sold", "Offers", "Bought", "Bids"]:
        ws[f'{L(int_col)}{row}'].number_format = '#,##0'
    
    # Add formulas
    ws[f'{L("Overcut (g)")}{row}'].value = f'={L("Buy (g.s)")}{row}*{L("Overcut (%)")}{row}'
    ws[f'{L("Undercut (g)")}{row}'].value = f'={L("Sell (g.s)")}{row}*{L("Undercut (%)")}{row}'
    ws[f'{L("Theoretical Profit")}{row}'].value = f'=({L("Undercut (g)")}{row}*0.85)-{L("Overcut (g)")}{row}'
    ws[f'{L("Amount Received")}{row}'].value = f'={L("Undercut (g)")}{row}*0.85'
    ws[f'{L("Demand-Supply Gap (%)")}{row}'].value = f'=IF({L("Supply")}{row}=0,"",({L("Demand")}{row}-{L("Supply")}{row})/{L("Supply")}{row})'

# Add a standard filter dropdown to all columns
ws.auto_filter.ref = ws.dimensions

# Auto-fit column widths
for col_cells in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col_cells[0].column)
    for cell in col_cells:
        try:
            if cell.value:
                if isinstance(cell.value, bool):
                     max_length = max(max_length, len(str(ws[f"{col_letter}1"].value)))
                else:
                     max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(output_file)
print(f"Final workbook saved to {output_file}. Rows: {max_row - 1}. Formatting complete.")