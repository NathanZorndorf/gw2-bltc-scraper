
import matplotlib
matplotlib.use('Agg')
import os
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta, timezone
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
import time
from tzlocal import get_localzone
import argparse

# Constants
BASE_URL = "https://www.gw2bltc.com/en/tp/search"
DATAWARS_API_URL = "https://api.datawars2.ie/gw2/v2/history/json"
DEFAULT_PARAMS = {
    "profit-min": 500,
    "profit-pct-min": 10,
    "profit-pct-max": 100,
    "sold-day-min": 5,
    "bought-day-min": 5,
    "ipg": 200,
    "sort": "profit-pct",
    "page": 1
}
OVERCUT_PCT_DEFAULT = 1.10
UNDERCUT_PCT_DEFAULT = 0.90
ROI_TARGET_DEFAULT = 0.10

# get timezone
try:
    local_tz = get_localzone()
except Exception:
    local_tz = timezone.utc

# --------------------
# HELPERS
# --------------------
def parse_gold_silver(td):
    gold = silver = 0
    for span in td.find_all("span"):
        classes = span.get("class", [])
        if "cur-t1c" in classes:
            gold = int(span.get_text(strip=True).replace(",", "") or 0)
        elif "cur-t1b" in classes:
            silver = int(span.get_text(strip=True) or 0)
    return round(gold + silver / 100, 2)

def parse_int(td):
    txt = td.get_text(strip=True).replace(",", "")
    try:
        return int(txt)
    except ValueError:
        return 0
    
def get_datawars_data(item_ids, status_callback, days=7):
    """Fetches and processes data from the DataWars2 API for multiple item IDs with retry logic."""
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=days)
    params = {
        "itemID": ",".join(item_ids),
        "start": start_date.strftime('%Y-%m-%dT%H:%M:%SZ'),
        "end": end_date.strftime('%Y-%m-%dT%H:%M:%SZ')
    }

    retries = 3
    backoff_factor = 0.5
    for i in range(retries):
        try:
            status_callback(f"Fetching DataWars2 data for items: {item_ids}")
            r = requests.get(DATAWARS_API_URL, params=params, timeout=10)
            r.raise_for_status()
            data = r.json()
            if not data:
                return {}

            # Verification check
            returned_item_ids = {str(d['itemID']) for d in data}
            if len(returned_item_ids) != len(item_ids):
                status_callback(f"Warning: Requested {len(item_ids)} items, but received data for {len(returned_item_ids)}.")

            results = {}
            for item_id in item_ids:
                item_data = [d for d in data if str(d['itemID']) == item_id]
                if not item_data:
                    results[item_id] = None
                    continue

                df = pd.DataFrame(item_data)
                required_cols = ['buy_price_avg','sell_price_avg','buy_price_max', 'sell_price_min', 'buy_listed', 'buy_sold', 'sell_listed', 'sell_sold', 'buy_quantity', 'sell_quantity']
                for col in required_cols:
                    if col not in df.columns:
                        df[col] = 0
                for col in required_cols:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

                if df.empty or df['buy_price_avg'].sum() == 0 or df['sell_price_avg'].sum() == 0:
                    results[item_id] = None
                    continue
                
                # New calculations
                avg_buy_price = df[df['buy_price_avg'] > 0]['buy_price_avg'].mean() / 10000
                avg_sell_price = df[df['sell_price_avg'] > 0]['sell_price_avg'].mean() / 10000
                std_dev_buy_price = df[df['buy_price_avg'] > 0]['buy_price_avg'].std() / 10000
                std_dev_sell_price = df[df['sell_price_avg'] > 0]['sell_price_avg'].std() / 10000

                # Instantaneous prices
                buy_price_inst = (df['buy_price_max'].iloc[-1] / 10000) if not df.empty else 0
                sell_price_inst = (df['sell_price_min'].iloc[-1] / 10000) if not df.empty else 0

                # New columns
                cov_buy = std_dev_buy_price / avg_buy_price if avg_buy_price else 0
                cov_sell = std_dev_sell_price / avg_sell_price if avg_sell_price else 0
                iv_buy = (buy_price_inst - avg_buy_price) / avg_buy_price if avg_buy_price else 0
                iv_sell = (sell_price_inst - avg_sell_price) / avg_sell_price if avg_sell_price else 0

                results[item_id] = {
                    "Buy Price (Inst.)": buy_price_inst,
                    "Sell Price (Inst.)": sell_price_inst,
                    "Demand": int(df['buy_quantity'].mean()),
                    "Supply": int(df['sell_quantity'].mean()),
                    "Bought": int(df['buy_sold'].sum()),
                    "Sold": int(df['sell_sold'].sum()),
                    "Bids": int(df['buy_listed'].sum()),
                    "Offers": int(df['sell_listed'].sum()),
                    "Avg Buy Price": avg_buy_price,
                    "Avg Sell Price": avg_sell_price,
                    "Std Dev Buy Price": std_dev_buy_price,
                    "Std Dev Sell Price": std_dev_sell_price,
                    "Coefficient of Variation (Buy)": cov_buy,
                    "Coefficient of Variation (Sell)": cov_sell,
                    "Instantaneous Volatility (Buy)": iv_buy,
                    "Instantaneous Volatility (Sell)": iv_sell,
                }
            return results
        except requests.exceptions.RequestException as e:
            status_callback(f"Failed to get data for items {item_ids}: {e}. Retrying ({i+1}/{retries})...")
            time.sleep(backoff_factor * (2 ** i))
        except (ValueError, KeyError) as e:
            status_callback(f"Failed to parse data for items {item_ids}: {e}")
            return {}
    status_callback(f"Failed to get data for items {item_ids} after {retries} retries.")
    return {}


def run_scraper(historical: bool, output_dir: str, days: int = 7, pages: int = 0, status_callback=None):
    if status_callback is None:
        status_callback = print

    os.makedirs(output_dir, exist_ok=True)
    input_file = os.path.join(output_dir, "scraper-results.xlsx")
    output_file = os.path.join(output_dir, "scraper-results-new.xlsx")

    status_callback(f"Your local timezone is: {local_tz}")

    all_rows = []
    params = DEFAULT_PARAMS.copy()
    scrape_time_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    while True:
        if pages > 0 and params['page'] > pages:
            status_callback(f"Reached page limit of {pages}.")
            break
        status_callback(f"Fetching page {params['page']}...")
        try:
            r = requests.get(BASE_URL, params=params, timeout=20)
            r.raise_for_status()
        except requests.exceptions.RequestException as e:
            status_callback(f"Request failed: {e}")
            break

        soup = BeautifulSoup(r.text, "html.parser")
        rows = soup.select("table.table-result tr")[1:]
        if not rows:
            status_callback("No more pages found.")
            break

        page_items = []
        page_item_ids = []
        for row in rows:
            cols = row.find_all("td")
            if len(cols) < 12: continue
            item_name = cols[1].get_text(strip=True)
            link_tag = cols[1].find("a", href=True)
            item_link = f"https://www.gw2bltc.com{link_tag['href']}" if link_tag else ""
            if not item_link: continue
            item_id = item_link.split('/')[-1].split('-')[0]
            item_data = {
                "item_id": item_id, "item_name": item_name, "item_link": item_link,
                "Buy Price (Inst.)": parse_gold_silver(cols[3]), "Sell Price (Inst.)": parse_gold_silver(cols[2]),
                "Demand": parse_int(cols[7]), "Supply": parse_int(cols[6]),
                "Bought": parse_int(cols[10]), "Sold": parse_int(cols[8]),
                "Bids": parse_int(cols[11]), "Offers": parse_int(cols[9])
            }
            page_items.append(item_data)
            page_item_ids.append(item_id)

        for i in range(0, len(page_item_ids), 50):
            batch_ids = page_item_ids[i:i+50]
            batch_items = page_items[i:i+50]
            api_data_dict = get_datawars_data(batch_ids, status_callback, days=days) if historical else {item["item_id"]: None for item in batch_items}

            for item_data in batch_items:
                api_data = api_data_dict.get(item_data["item_id"])
                row_data = [
                    item_data["item_name"], item_data["item_link"], scrape_time_str,
                    item_data["Buy Price (Inst.)"], item_data["Sell Price (Inst.)"],
                    item_data['Demand'], item_data['Supply'], item_data['Bought'], item_data['Sold'],
                    item_data['Bids'], item_data['Offers']
                ]
                if api_data:
                    row_data.extend([
                        api_data["Avg Buy Price"], api_data["Avg Sell Price"],
                        api_data["Std Dev Buy Price"], api_data["Std Dev Sell Price"],
                        api_data["Coefficient of Variation (Buy)"], api_data["Coefficient of Variation (Sell)"],
                        api_data["Instantaneous Volatility (Buy)"], api_data["Instantaneous Volatility (Sell)"]
                    ])
                else:
                    row_data.extend(['', '', '', '', '', '', '', ''])
                all_rows.append(row_data)
        params["page"] += 1

    if not all_rows:
        status_callback("No data scraped.")
        return

    status_callback("Scraping complete. Processing data...")

    if os.path.exists(input_file):
        try:
            existing_df = pd.read_excel(input_file, sheet_name='scraper-results')
            existing_df = existing_df[existing_df["Buy Order Placed"] == True]
        except Exception as e:
            status_callback(f"Could not read existing file {input_file}: {e}")
            existing_df = pd.DataFrame()
    else:
        existing_df = pd.DataFrame()

    df = pd.DataFrame(all_rows, columns=[
        "Item Name", "Item Link", "Date of Scrape", "Buy Price (Inst.)", "Sell Price (Inst.)",
        "Demand", "Supply", "Bought", "Sold", "Bids", "Offers",
        "Avg Buy Price", "Avg Sell Price", "Std Dev Buy Price", "Std Dev Sell Price",
        "Coefficient of Variation (Buy)", "Coefficient of Variation (Sell)",
        "Instantaneous Volatility (Buy)", "Instantaneous Volatility (Sell)"
    ])

    final_column_order = [
        "Item Name", "Item Link", "Date of Scrape", "Buy Price (Inst.)", "Sell Price (Inst.)",
        "Demand", "Supply", "Bought", "Sold", "Bids", "Offers",
        "Avg Buy Price", "Avg Sell Price", "Std Dev Buy Price", "Std Dev Sell Price",
        "Coefficient of Variation (Buy)", "Coefficient of Variation (Sell)",
        "Instantaneous Volatility (Buy)", "Instantaneous Volatility (Sell)",
        "Overcut (%)", "Undercut (%)", "Overcut (g)", "Undercut (g)",
        "Max Flips / Day", "Bought/Bids", "Sold/Offers",
        "Buy-Through Rate (%)", "Sell-Through Rate (%)", "Flip-Through Rate (%)",
        "Optimal Qty", "Dynamic Sell-Through Rate (%)", "E(Sales | Q = Optimal Q)",
        "E(Profit | Q = Optimal Q)", "Optimal Investment (g)", "E(ROI | Q = Optimal Q)", "Time to Sell (Q Optimal)",
        "Target ROI", "Optimal Buy Price | Target ROI", "Optimal Qty | Target ROI",
        "Theoretical Return | Target ROI",
        "Actual Qty Ordered", "Actual Buy Price", "Actual Sell Price",
        "Buy Order Placed", "Sell Order Placed", "Sold (manual)"
    ]

    for col in final_column_order:
        if col not in df.columns:
            df[col] = ""

    df["Overcut (%)"] = OVERCUT_PCT_DEFAULT
    df["Undercut (%)"] = UNDERCUT_PCT_DEFAULT
    df["Target ROI"] = ROI_TARGET_DEFAULT
    df["Buy Order Placed"] = False
    df["Sell Order Placed"] = False
    df["Sold (manual)"] = False

    for col in final_column_order:
        if col not in existing_df.columns:
            existing_df[col] = ""

    combined_df = pd.concat([existing_df[final_column_order], df[final_column_order]], ignore_index=True)
    combined_df.to_excel(output_file, index=False)

    status_callback("Formatting Excel file...")

    wb = load_workbook(output_file)
    ws = wb.active
    ws.title = "scraper-results"
    ws.freeze_panes = 'B2'

    header_to_idx = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1)}
    def L(name): return get_column_letter(header_to_idx.get(name))
    max_row = ws.max_row

    for row in range(2, max_row + 1):
        # Number formats
        ws[f'{L("Buy Price (Inst.)")}{row}'].number_format = '0.00'
        ws[f'{L("Sell Price (Inst.)")}{row}'].number_format = '0.00'
        ws[f'{L("Avg Buy Price")}{row}'].number_format = '0.00'
        ws[f'{L("Avg Sell Price")}{row}'].number_format = '0.00'
        ws[f'{L("Std Dev Buy Price")}{row}'].number_format = '0.00'
        ws[f'{L("Std Dev Sell Price")}{row}'].number_format = '0.00'
        ws[f'{L("Coefficient of Variation (Buy)")}{row}'].number_format = '0%'
        ws[f'{L("Coefficient of Variation (Sell)")}{row}'].number_format = '0%'
        ws[f'{L("Instantaneous Volatility (Buy)")}{row}'].number_format = '0%'
        ws[f'{L("Instantaneous Volatility (Sell)")}{row}'].number_format = '0%'
        ws[f'{L("Overcut (%)")}{row}'].number_format = '0%'
        ws[f'{L("Undercut (%)")}{row}'].number_format = '0%'
        ws[f'{L("Overcut (g)")}{row}'].number_format = '0.00'
        ws[f'{L("Undercut (g)")}{row}'].number_format = '0.00'
        ws[f'{L("Max Flips / Day")}{row}'].number_format = '0'
        ws[f'{L("Bought/Bids")}{row}'].number_format = '0.00'
        ws[f'{L("Sold/Offers")}{row}'].number_format = '0.00'
        ws[f'{L("Buy-Through Rate (%)")}{row}'].number_format = '0%'
        ws[f'{L("Sell-Through Rate (%)")}{row}'].number_format = '0%'
        ws[f'{L("Flip-Through Rate (%)")}{row}'].number_format = '0%'
        ws[f'{L("Optimal Qty")}{row}'].number_format = '0'
        ws[f'{L("Dynamic Sell-Through Rate (%)")}{row}'].number_format = '0%'
        ws[f'{L("E(Sales | Q = Optimal Q)")}{row}'].number_format = '0.00'
        ws[f'{L("E(Profit | Q = Optimal Q)")}{row}'].number_format = '0.00'
        ws[f'{L("Optimal Investment (g)")}{row}'].number_format = '0.00'
        ws[f'{L("E(ROI | Q = Optimal Q)")}{row}'].number_format = '0%'
        ws[f'{L("Time to Sell (Q Optimal)")}{row}'].number_format = '0.00'
        ws[f'{L("Actual Qty Ordered")}{row}'].number_format = '0'
        ws[f'{L("Actual Buy Price")}{row}'].number_format = '0.00'

        # Formatting for new columns
        ws[f'{L("Target ROI")}{row}'].number_format = '0%'
        ws[f'{L("Optimal Qty | Target ROI")}{row}'].number_format = '0'
        ws[f'{L("Optimal Buy Price | Target ROI")}{row}'].number_format = '0.00'
        ws[f'{L("Theoretical Return | Target ROI")}{row}'].number_format = '0.00'

        for int_col in ["Demand", "Supply", "Bought", "Sold", "Bids", "Offers"]:
            ws[f'{L(int_col)}{row}'].number_format = '#,##0'

        # Formulas
        ws[f'{L("Overcut (g)")}{row}'].value = f'={L("Buy Price (Inst.)")}{row}*{L("Overcut (%)")}{row}'
        ws[f'{L("Undercut (g)")}{row}'].value = f'={L("Sell Price (Inst.)")}{row}*{L("Undercut (%)")}{row}'
        ws[f'{L("Max Flips / Day")}{row}'].value = f'=MIN({L("Bought")}{row},{L("Sold")}{row})'
        ws[f'{L("Bought/Bids")}{row}'].value = f'=IFERROR({L("Bought")}{row}/{L("Bids")}{row},"")'
        ws[f'{L("Sold/Offers")}{row}'].value = f'=IFERROR({L("Sold")}{row}/{L("Offers")}{row},"")'
        ws[f'{L("Buy-Through Rate (%)")}{row}'].value = f'=IF({L("Bids")}{row}=0,IF({L("Bought")}{row}>0,1,0),MIN(1,{L("Bought")}{row}/{L("Bids")}{row}))'
        ws[f'{L("Sell-Through Rate (%)")}{row}'].value = f'=IF({L("Offers")}{row}=0,IF({L("Sold")}{row}>0,1,0),MIN(1,{L("Sold")}{row}/{L("Offers")}{row}))'
        ws[f'{L("Flip-Through Rate (%)")}{row}'].value = f'={L("Buy-Through Rate (%)")}{row}*{L("Sell-Through Rate (%)")}{row}'
        ws[f'{L("Optimal Qty")}{row}'].value = f'=LET(q,ROUND(SQRT({L("Sold")}{row}*{L("Offers")}{row}*{L("Undercut (g)")}{row}*0.85/{L("Overcut (g)")}{row})-{L("Offers")}{row}),IF(q<0,0,MIN(q,{L("Max Flips / Day")}{row})))'
        ws[f'{L("Dynamic Sell-Through Rate (%)")}{row}'].value = f'=IFERROR(IF({L("Optimal Qty")}{row}>0,MIN(1,{L("Sold")}{row}/({L("Offers")}{row}+{L("Optimal Qty")}{row})),NA()),"")'
        ws[f'{L("E(Sales | Q = Optimal Q)")}{row}'].value = f'={L("Optimal Qty")}{row}*{L("Dynamic Sell-Through Rate (%)")}{row}'
        ws[f'{L("E(Profit | Q = Optimal Q)")}{row}'].value = f'={L("E(Sales | Q = Optimal Q)")}{row}*{L("Undercut (g)")}{row}*0.85-{L("Overcut (g)")}{row}*{L("Optimal Qty")}{row}'
        ws[f'{L("Optimal Investment (g)")}{row}'].value = f'={L("Optimal Qty")}{row}*{L("Overcut (g)")}{row}'
        ws[f'{L("E(ROI | Q = Optimal Q)")}{row}'].value = f'=IFERROR({L("E(Profit | Q = Optimal Q)")}{row}/{L("Optimal Investment (g)")}{row},0)'
        ws[f'{L("Time to Sell (Q Optimal)")}{row}'].value = f'=({L("Offers")}{row} + {L("Optimal Qty")}{row})/{L("Sold")}{row}'

        # New formulas for Target ROI
        ws[f'{L("Target ROI")}{row}'].value = f'={ROI_TARGET_DEFAULT}'
        ws[f'{L("Optimal Buy Price | Target ROI")}{row}'].value = f'=LET(q, IFERROR(({L("Undercut (g)")}{row}*0.85)/(1+{L("Target ROI")}{row}), 0), IF(q > {L("Buy Price (Inst.)")}{row}, q, 0)))'
        ws[f'{L("Optimal Qty | Target ROI")}{row}'].value = f'=LET(q,ROUND(SQRT({L("Sold")}{row}*{L("Offers")}{row}*{L("Undercut (g)")}{row}*0.85/{L("Optimal Buy Price | Target ROI")}{row}) - {L("Offers")}{row}), IF(q<0,0,MIN(q,{L("Max Flips / Day")}{row})))'
        ws[f'{L("Theoretical Return | Target ROI")}{row}'].value = f'=({L("Undercut (g)")}{row}*0.85-{L("Optimal Buy Price | Target ROI")}{row})*{L("Optimal Qty | Target ROI")}{row}'

    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(8, max_length + 2)

    wb.save(output_file)
    status_callback(f"Success! Final workbook saved to {output_file}.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GW2 BLTC Scraper")
    parser.add_argument('--historical', action='store_true', help='Query DataWars2 API for historical data')
    parser.add_argument('--output_dir', type=str, default='.', help='Directory to save the output file')
    parser.add_argument('--days', type=int, default=7, help='Number of days of historical data to query')
    parser.add_argument('--pages', type=int, default=0, help='Number of pages to scrape (0 for all)')
    args = parser.parse_args()

    run_scraper(historical=args.historical, output_dir=args.output_dir, days=args.days, pages=args.pages)