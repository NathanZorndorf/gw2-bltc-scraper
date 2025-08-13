import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# Import classes needed for applying filters
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters
)

# --------------------
# CONFIG
# --------------------
BASE_URL = "https://www.gw2bltc.com/en/tp/search"
PARAMS = {
    # "profit-min": 0,
    # "profit-pct-min": 0,
    # "profit-pct-max": 100,
    "sold-day-min": 10,
    "bought-day-min": 10,
    "ipg": 200,
    "sort": "profit-pct",
    "page": 1
}

OVERCUT_PCT_DEFAULT = 1.10
UNDERCUT_PCT_DEFAULT = 0.90

output_file = "gw2_trading_post_filtered.xlsx"

# --------------------
# SCRAPE -> list of rows
# --------------------
def parse_gold_silver(td):
    """Helper function to parse gold and silver values from table cells."""
    gold = silver = 0
    for span in td.find_all("span"):
        classes = span.get("class", [])
        if "cur-t1c" in classes:
            try:
                gold = int(span.get_text(strip=True))
            except ValueError:
                gold = 0
        elif "cur-t1b" in classes:
            try:
                silver = int(span.get_text(strip=True))
            except ValueError:
                silver = 0
    return round(gold + silver / 100, 2)

def parse_int(td):
    """Helper function to parse integer values from table cells."""
    txt = td.get_text(strip=True).replace(",", "")
    try:
        return int(txt)
    except ValueError:
        return 0

all_rows = []
scrape_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Loop through pages on the website until there's no more data
while True:
    print(f"Fetching page {PARAMS['page']}...")
    try:
        r = requests.get(BASE_URL, params=PARAMS, timeout=20)
        r.raise_for_status()  # Raise an exception for bad status codes (4xx or 5xx)
    except requests.exceptions.RequestException as e:
        print(f"Request failed: {e}")
        break

    soup = BeautifulSoup(r.text, "html.parser")
    rows = soup.select("table.table-result tr")[1:]  # skip header

    if not rows:
        print("No more rows found on page. Ending scrape.")
        break

    page_has_data = False
    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 12:
            continue

        page_has_data = True
        item_name = cols[1].get_text(strip=True)
        link_tag = cols[1].find("a", href=True)
        item_link = f"https://www.gw2bltc.com{link_tag['href']}" if link_tag else ""

        sell = parse_gold_silver(cols[2])
        buy = parse_gold_silver(cols[3])
        supply = parse_int(cols[6])
        demand = parse_int(cols[7])
        sold = parse_int(cols[8])
        offers = parse_int(cols[9])
        bought = parse_int(cols[10])
        bids = parse_int(cols[11])

        all_rows.append([
            item_name, item_link, scrape_time, sell, buy,
            supply, demand, sold, offers, bought, bids
        ])

    if not page_has_data:
        print("Page had no valid data rows. Ending scrape.")
        break

    PARAMS["page"] += 1

# --------------------
# Build and Reorder DataFrame
# --------------------
df = pd.DataFrame(all_rows, columns=[
    "Item Name", "Item Link", "Date of Scrape",
    "Sell (g.s)", "Buy (g.s)", "Supply", "Demand", "Sold",
    "Offers", "Bought", "Bids"
])

# Add placeholder columns that will either be static or filled by formulas later
df["Overcut (%)"] = OVERCUT_PCT_DEFAULT
df["Undercut (%)"] = UNDERCUT_PCT_DEFAULT
df["Overcut (g)"] = None
df["Undercut (g)"] = None
df["Theoretical Profit"] = None
df["Amount Received"] = None
df["Demand-Supply Gap (%)"] = None
df["Order Placed"] = False
df["Order Successful"] = False
df["Sold (manual)"] = False

# Reorder columns to the desired final layout
final_column_order = [
    # Scraped data
    "Item Name", "Item Link", "Date of Scrape", "Buy (g.s)", "Sell (g.s)",
    "Supply", "Demand", "Offers", "Sold", "Bids", "Bought",
    # Static per-row %
    "Overcut (%)", "Undercut (%)",
    # Calculated formulas
    "Overcut (g)", "Undercut (g)", "Theoretical Profit", "Amount Received", "Demand-Supply Gap (%)",
    # Manual user-entry
    "Order Placed", "Order Successful", "Sold (manual)"
]
df = df[final_column_order]

# Save the reordered dataframe to Excel. This file will be modified further.
df.to_excel(output_file, index=False)
print(f"Base workbook written to {output_file} with reordered columns.")

# --------------------
# Insert Formulas, Format, and Add Filters with OpenPyXL
# --------------------
wb = load_workbook(output_file)
ws = wb.active
max_row = ws.max_row

# Build a header->col_index map (1-based) to easily find columns by name
header_to_idx = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1)}

# Helper to get column letter from column name
def L(name):
    if name not in header_to_idx:
        raise KeyError(f"Column '{name}' not found in the Excel sheet headers.")
    return get_column_letter(header_to_idx[name])

# Get letters for columns needed in formulas
sell_col_letter = L("Sell (g.s)")
buy_col_letter = L("Buy (g.s)")
overcut_pct_letter = L("Overcut (%)")
undercut_pct_letter = L("Undercut (%)")
supply_letter = L("Supply")
demand_letter = L("Demand")
overcut_g_letter = L("Overcut (g)")
undercut_g_letter = L("Undercut (g)")

# Write formulas and apply number formats row-by-row
for row in range(2, max_row + 1):
    # --- Formulas ---
    ws[f'{L("Overcut (g)")}{row}'] = f"={buy_col_letter}{row}*{overcut_pct_letter}{row}"
    ws[f'{L("Undercut (g)")}{row}'] = f"={sell_col_letter}{row}*{undercut_pct_letter}{row}"
    ws[f'{L("Theoretical Profit")}{row}'] = f"={undercut_g_letter}{row}*0.85 - {overcut_g_letter}{row}"
    ws[f'{L("Amount Received")}{row}'] = f"={undercut_g_letter}{row}*0.85"
    ws[f'{L("Demand-Supply Gap (%)")}{row}'] = f"=IF({supply_letter}{row}=0, 0, ({demand_letter}{row}-{supply_letter}{row})/{supply_letter}{row})"

    # --- Number Formatting ---
    ws[f'{L("Buy (g.s)")}{row}'].number_format = '0.00'
    ws[f'{L("Sell (g.s)")}{row}'].number_format = '0.00'
    ws[f'{L("Overcut (%)")}{row}'].number_format = '0.00%'
    ws[f'{L("Undercut (%)")}{row}'].number_format = '0.00%'
    ws[f'{L("Overcut (g)")}{row}'].number_format = '0.00'
    ws[f'{L("Undercut (g)")}{row}'].number_format = '0.00'
    ws[f'{L("Theoretical Profit")}{row}'].number_format = '0.00'
    ws[f'{L("Amount Received")}{row}'].number_format = '0.00'
    ws[f'{L("Demand-Supply Gap (%)")}{row}'].number_format = '0.00%'

    for int_col in ["Supply", "Demand", "Sold", "Offers", "Bought", "Bids"]:
        ws[f'{L(int_col)}{row}'].number_format = '#,##0'

# --- Apply AutoFilter ---
# Set the range for the filter (the entire data area)
ws.auto_filter.ref = ws.dimensions

# Filter 1: Theoretical Profit > 0.5
profit_col_idx = header_to_idx["Theoretical Profit"]
profit_filter_col = FilterColumn(colId=profit_col_idx - 1) # colId is 0-based
# --- FIX: Convert numeric value to string for the filter ---
profit_custom_filter = CustomFilter(operator="greaterThan", val="0.5")
profit_filter_col.customFilters = CustomFilters(customFilter=[profit_custom_filter])
ws.auto_filter.filterColumn.append(profit_filter_col)

# Filter 2: Demand-Supply Gap between -0.3 and 0.3
gap_col_idx = header_to_idx["Demand-Supply Gap (%)"]
gap_filter_col = FilterColumn(colId=gap_col_idx - 1) # colId is 0-based
# --- FIX: Convert numeric values to strings for the filter ---
gap_filter_1 = CustomFilter(operator="greaterThan", val="-0.3")
gap_filter_2 = CustomFilter(operator="lessThan", val="0.3")
gap_filter_col.customFilters = CustomFilters(customFilter=[gap_filter_1, gap_filter_2])
ws.auto_filter.filterColumn.append(gap_filter_col)

# --------------------
# Auto-fit column widths
# --------------------
for col_cells in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col_cells[0].column)
    for cell in col_cells:
        try:
            if cell.value:
                # Adjust length for boolean False, which is shorter than the header
                if isinstance(cell.value, bool):
                     max_length = max(max_length, len(str(ws[f"{col_letter}1"].value)))
                else:
                     max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(output_file)
print(f"Final workbook saved to {output_file}. Rows: {max_row - 1}. Filters applied.")
